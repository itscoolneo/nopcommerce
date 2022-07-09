import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)


def getCoulmnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,colnum).value

def writeData(file,sheetname,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)

def fillGreenColor(file,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greeFill = PatternFill(start_color='60b212',
                           end_color='60b212',
                           fill_type='solid')
    sheet.cell(rownum,colnum).fill=greeFill
    workbook.save(file)


def fillRedColor(file,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redFill = PatternFill(start_color='ff0000',
                           end_color='ff0000',
                           fill_type='solid')
    sheet.cell(rownum,colnum).fill= redFill
    workbook.save(file)